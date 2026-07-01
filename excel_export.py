import io
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── 색상 팔레트 ───────────────────────────────────────────────────
COLOR = {
    'header_bg':   'FF37474F',
    'header_font': 'FFFFFFFF',
    'pos':         'FF2E7D32',  # 증가 (초록)
    'neg':         'FFC62828',  # 감소 (빨강)
    'neutral':     'FF9E9E9E',
    'row_even':    'FFF5F5F5',
    'row_odd':     'FFFFFFFF',
    'border':      'FFDDDDDD',
}

_thin = Side(style='thin', color=COLOR['border'])
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_style(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=COLOR['header_font'], size=10)
    cell.fill = PatternFill('solid', fgColor=COLOR['header_bg'])
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _border


def _num_fmt(cell, value, fmt='#,##0'):
    cell.value = value
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal='right')
    cell.border = _border


def _change_style(cell, value):
    _num_fmt(cell, value, '+#,##0;-#,##0;0')
    if value is None or pd.isna(value):
        cell.value = None
        return
    if value > 0:
        cell.font = Font(color=COLOR['pos'], bold=True)
    elif value < 0:
        cell.font = Font(color=COLOR['neg'], bold=True)
    else:
        cell.font = Font(color=COLOR['neutral'])


def generate_excel(df: pd.DataFrame, campaigns: dict = None,
                   df_conv: pd.DataFrame = None,
                   df_adspend: pd.DataFrame = None,
                   df_content: pd.DataFrame = None,
                   rooms: dict = None) -> bytes:
    """전체 인원 데이터를 받아 포맷된 Excel 파일을 bytes로 반환."""
    wb = Workbook()

    _build_daily_sheet(wb, df, campaigns)
    _build_summary_sheet(wb, df, campaigns)
    _build_trend_sheet(wb, df)
    if df_conv is not None and not df_conv.empty:
        _build_conversion_sheet(wb, df_conv, campaigns or {})
    if df_adspend is not None and not df_adspend.empty:
        _build_adspend_sheet(wb, df_adspend, campaigns or {})
    if df_content is not None and not df_content.empty:
        _build_content_sheet(wb, df_content)
    if not df.empty:
        _build_ranking_sheet(wb, df, rooms or {})
    if df_adspend is not None and not df_adspend.empty and not df.empty:
        _build_cpm_sheet(wb, df, df_adspend, rooms or {})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 시트 1: 일별 상세 ──────────────────────────────────────────────

def _build_daily_sheet(wb, df, campaigns):
    ws = wb.active
    ws.title = '일별 상세'

    headers = ['날짜', '방 번호', '채팅방', '강의명', '상품', '총원', '전일', '증감']
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=col), h)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    sorted_df = df.sort_values(['date', 'room_num'], ascending=[False, True])

    for r_idx, (_, row) in enumerate(sorted_df.iterrows(), 2):
        rn = int(row['room_num'])
        camp = (campaigns or {}).get(rn, {})
        fill = PatternFill('solid', fgColor=COLOR['row_even'] if r_idx % 2 == 0 else COLOR['row_odd'])

        def cell(col):
            c = ws.cell(row=r_idx, column=col)
            c.fill = fill
            return c

        c = cell(1); c.value = str(row['date']); c.alignment = Alignment(horizontal='center'); c.border = _border
        c = cell(2); c.value = rn; c.alignment = Alignment(horizontal='center'); c.border = _border
        c = cell(3); c.value = str(row['room_name']); c.alignment = Alignment(horizontal='left'); c.border = _border
        c = cell(4); c.value = camp.get('campaign_name', '-'); c.alignment = Alignment(horizontal='left'); c.border = _border
        c = cell(5); c.value = camp.get('product', '-'); c.alignment = Alignment(horizontal='center'); c.border = _border
        _num_fmt(cell(6), int(row['members']))
        prev = row['prev_members']
        _num_fmt(cell(7), int(prev) if not pd.isna(prev) else None)
        chg = row['change']
        _change_style(cell(8), int(chg) if not pd.isna(chg) else None)

    col_widths = [12, 8, 20, 20, 10, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 시트 2: 날짜별 요약 ────────────────────────────────────────────

def _build_summary_sheet(wb, df, campaigns):
    ws = wb.create_sheet('날짜별 요약')

    if df.empty:
        ws.cell(row=1, column=1).value = '데이터 없음'
        return

    # 날짜별 총원 합계·증감 합계
    summary = (
        df.groupby('date')
        .agg(total_members=('members', 'sum'), net_change=('change', 'sum'))
        .reset_index()
        .sort_values('date', ascending=False)
    )

    headers = ['날짜', '전체 총원', '전일 대비 순증감']
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=col), h)

    ws.freeze_panes = 'A2'

    for r_idx, (_, row) in enumerate(summary.iterrows(), 2):
        ws.cell(row=r_idx, column=1).value = str(row['date'])
        ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=r_idx, column=1).border = _border
        _num_fmt(ws.cell(row=r_idx, column=2), int(row['total_members']))
        chg = row['net_change']
        _change_style(ws.cell(row=r_idx, column=3), int(chg) if not pd.isna(chg) else None)

    for i, w in enumerate([12, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 막대 차트 삽입
    if len(summary) >= 2:
        chart = BarChart()
        chart.title = '날짜별 전체 총원'
        chart.style = 10
        chart.height = 12
        chart.width = 20

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(summary) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(summary) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, 'E2')


# ── 시트 3: 채팅방별 추이 ─────────────────────────────────────────

def _build_trend_sheet(wb, df):
    ws = wb.create_sheet('채팅방별 추이')

    if df.empty:
        ws.cell(row=1, column=1).value = '데이터 없음'
        return

    # 날짜 × 채팅방 피벗
    pivot = df.pivot_table(index='date', columns='room_num', values='members', aggfunc='first')
    pivot = pivot.sort_index(ascending=False)
    pivot.reset_index(inplace=True)

    # 헤더
    _header_style(ws.cell(row=1, column=1), '날짜')
    for col_idx, col_name in enumerate(pivot.columns[1:], 2):
        _header_style(ws.cell(row=1, column=col_idx), f'채팅방 {int(col_name)}')

    ws.freeze_panes = 'A2'

    for r_idx, (_, row) in enumerate(pivot.iterrows(), 2):
        ws.cell(row=r_idx, column=1).value = str(row['date'])
        ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=r_idx, column=1).border = _border
        for col_idx, col_name in enumerate(pivot.columns[1:], 2):
            val = row[col_name]
            _num_fmt(ws.cell(row=r_idx, column=col_idx),
                     int(val) if not pd.isna(val) else None)

    for i in range(1, len(pivot.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12


# ── 시트 4: 전환 분석 ─────────────────────────────────────────────

def _build_conversion_sheet(wb, df_conv: pd.DataFrame, campaigns: dict):
    ws = wb.create_sheet('전환 분석')

    if df_conv.empty:
        ws.cell(row=1, column=1).value = '전환 데이터 없음'
        return

    headers = ['날짜', '강의명', '상품', '기수', '신청자', '수강확정', '전환율(%)', '매출(원)', '메모']
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    sorted_df = df_conv.sort_values(['date', 'room_num'], ascending=[False, True])

    for r_idx, (_, row) in enumerate(sorted_df.iterrows(), 2):
        rn   = int(row['room_num'])
        camp = campaigns.get(rn, {})
        appl = int(row['applicants'])
        conf = int(row['confirmed'])
        conv_pct = round(conf / appl * 100, 1) if appl > 0 else 0
        fill = PatternFill('solid', fgColor=COLOR['row_even'] if r_idx % 2 == 0 else COLOR['row_odd'])

        def cell(col):
            c = ws.cell(row=r_idx, column=col)
            c.fill = fill
            c.border = _border
            return c

        c = cell(1); c.value = str(row['date']); c.alignment = Alignment(horizontal='center')
        c = cell(2); c.value = camp.get('campaign_name', f'채팅방 {rn}'); c.alignment = Alignment(horizontal='left')
        c = cell(3); c.value = camp.get('product', '-'); c.alignment = Alignment(horizontal='center')
        c = cell(4); c.value = camp.get('cohort', '-'); c.alignment = Alignment(horizontal='center')
        _num_fmt(cell(5), appl)
        _num_fmt(cell(6), conf)
        c7 = cell(7)
        c7.value = conv_pct
        c7.number_format = '0.0'
        c7.alignment = Alignment(horizontal='center')
        if conv_pct >= 80:
            c7.font = Font(bold=True, color=COLOR['pos'])
        _num_fmt(cell(8), int(row['revenue']))
        c = cell(9); c.value = str(row.get('memo', '') or ''); c.alignment = Alignment(horizontal='left')

    for i, w in enumerate([12, 22, 10, 8, 10, 10, 12, 14, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 시트 5: 광고비 ────────────────────────────────────────────────

def _build_adspend_sheet(wb, df_adspend: pd.DataFrame, campaigns: dict = None):
    ws = wb.create_sheet('광고비')

    if df_adspend.empty:
        ws.cell(row=1, column=1).value = '광고비 데이터 없음'
        return

    headers = ['날짜', '강의명', '채널', '광고비(원)', '노출수', '클릭수', 'CPC(원)', '메모']
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    sorted_df = df_adspend.sort_values(['date', 'room_num'], ascending=[False, True])

    for r_idx, (_, row) in enumerate(sorted_df.iterrows(), 2):
        rn     = int(row['room_num'])
        camp   = (campaigns or {}).get(rn, {})
        spend  = int(row['spend'])
        clicks = int(row.get('clicks', 0) or 0)
        imps   = int(row.get('impressions', 0) or 0)
        cpc    = round(spend / clicks) if clicks > 0 else 0
        fill = PatternFill('solid', fgColor=COLOR['row_even'] if r_idx % 2 == 0 else COLOR['row_odd'])

        def cell(col):
            c = ws.cell(row=r_idx, column=col)
            c.fill = fill
            c.border = _border
            return c

        c = cell(1); c.value = str(row['date']); c.alignment = Alignment(horizontal='center')
        c = cell(2); c.value = camp.get('campaign_name', f'채팅방 {rn}'); c.alignment = Alignment(horizontal='left')
        c = cell(3); c.value = str(row.get('channel', '-')); c.alignment = Alignment(horizontal='center')
        _num_fmt(cell(4), spend)
        _num_fmt(cell(5), imps)
        _num_fmt(cell(6), clicks)
        _num_fmt(cell(7), cpc)
        c = cell(8); c.value = str(row.get('memo', '') or ''); c.alignment = Alignment(horizontal='left')

    for i, w in enumerate([12, 22, 14, 14, 12, 10, 12, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 시트 6: 콘텐츠 기록 ──────────────────────────────────────────

def _build_content_sheet(wb, df_content: pd.DataFrame):
    ws = wb.create_sheet('콘텐츠 기록')

    if df_content.empty:
        ws.cell(row=1, column=1).value = '콘텐츠 데이터 없음'
        return

    headers = ['날짜', '채널', '유형', '제목', 'URL', '메모']
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for r_idx, (_, row) in enumerate(df_content.sort_values('date', ascending=False).iterrows(), 2):
        fill = PatternFill('solid', fgColor=COLOR['row_even'] if r_idx % 2 == 0 else COLOR['row_odd'])

        def cell(col):
            c = ws.cell(row=r_idx, column=col)
            c.fill = fill
            c.border = _border
            return c

        c = cell(1); c.value = str(row['date']); c.alignment = Alignment(horizontal='center')
        c = cell(2); c.value = str(row.get('channel', '-')); c.alignment = Alignment(horizontal='center')
        c = cell(3); c.value = str(row.get('content_type', '-')); c.alignment = Alignment(horizontal='center')
        c = cell(4); c.value = str(row.get('title', '-')); c.alignment = Alignment(horizontal='left')
        c = cell(5); c.value = str(row.get('url', '') or ''); c.alignment = Alignment(horizontal='left')
        c = cell(6); c.value = str(row.get('memo', '') or ''); c.alignment = Alignment(horizontal='left')

    for i, w in enumerate([12, 14, 14, 30, 40, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 시트 7: 주간 랭킹 ────────────────────────────────────────────

def _build_ranking_sheet(wb, df: pd.DataFrame, rooms: dict):
    ws = wb.create_sheet('주간 랭킹')
    ws.cell(row=1, column=1).value = f'기준일: {date.today()}'
    ws.cell(row=1, column=1).font = Font(bold=True)

    dates = sorted(df['date'].unique())
    if len(dates) < 2:
        ws.cell(row=2, column=1).value = '데이터 부족 (2일 이상 필요)'
        return

    # 5~9일 전 vs 최근 비교
    recent_end = dates[-1]
    lookback = min(9, len(dates) - 1)
    lookback = max(5, lookback)
    past_start = dates[-(lookback + 1)]

    recent = df[df['date'] == recent_end].set_index('room_num')['members']
    past   = df[df['date'] == past_start].set_index('room_num')['members']
    common = set(recent.index) & set(past.index)

    rows = []
    for rn in common:
        diff = int(recent[rn]) - int(past[rn])
        rows.append({'방 번호': rn, '채팅방': rooms.get(rn, f'채팅방 {rn}'),
                     f'{past_start} 인원': int(past[rn]),
                     f'{recent_end} 인원': int(recent[rn]),
                     '증감': diff})
    if not rows:
        ws.cell(row=2, column=1).value = '비교 가능한 데이터 없음'
        return

    result_df = pd.DataFrame(rows).sort_values('증감', ascending=False)
    headers = list(result_df.columns)
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=2, column=col), h)

    for r_idx, (_, row) in enumerate(result_df.iterrows(), 3):
        fill = PatternFill('solid', fgColor=COLOR['row_even'] if r_idx % 2 == 0 else COLOR['row_odd'])
        for col, key in enumerate(headers, 1):
            c = ws.cell(row=r_idx, column=col)
            c.fill = fill
            c.border = _border
            if key in ('증감',):
                _change_style(c, row[key])
            elif isinstance(row[key], (int, float)):
                _num_fmt(c, row[key])
            else:
                c.value = row[key]
                c.alignment = Alignment(horizontal='left')

    for i, w in enumerate([8, 22, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 시트 8: CPM 분석 ─────────────────────────────────────────────

def _build_cpm_sheet(wb, df: pd.DataFrame, df_adspend: pd.DataFrame, rooms: dict):
    ws = wb.create_sheet('CPM 분석')
    ws.cell(row=1, column=1).value = f'광고비 ÷ 인원증가 분석 (기준일: {date.today()})'
    ws.cell(row=1, column=1).font = Font(bold=True)

    dates = sorted(df['date'].unique())
    if len(dates) < 2:
        ws.cell(row=2, column=1).value = '데이터 부족'
        return

    first_date = dates[0]
    last_date  = dates[-1]
    first = df[df['date'] == first_date].set_index('room_num')['members']
    last  = df[df['date'] == last_date].set_index('room_num')['members']

    total_spend_by_room = df_adspend.groupby('room_num')['spend'].sum()

    rows = []
    for rn in set(last.index) & set(first.index):
        spend = int(total_spend_by_room.get(rn, 0))
        increase = int(last[rn]) - int(first[rn])
        if increase <= 0 or spend <= 0:
            continue
        cpm = round(spend / increase)
        rows.append({'방 번호': rn, '채팅방': rooms.get(rn, f'채팅방 {rn}'),
                     '총 광고비': spend, '인원 증가': increase, 'CPM(원/명)': cpm})

    if not rows:
        ws.cell(row=2, column=1).value = '광고비 + 인원 증가 데이터 없음'
        return

    result_df = pd.DataFrame(rows).sort_values('CPM(원/명)')
    headers = list(result_df.columns)
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=2, column=col), h)

    for r_idx, (_, row) in enumerate(result_df.iterrows(), 3):
        fill = PatternFill('solid', fgColor=COLOR['row_even'] if r_idx % 2 == 0 else COLOR['row_odd'])
        for col, key in enumerate(headers, 1):
            c = ws.cell(row=r_idx, column=col)
            c.fill = fill
            c.border = _border
            if isinstance(row[key], (int, float)):
                _num_fmt(c, row[key])
            else:
                c.value = row[key]
                c.alignment = Alignment(horizontal='left')

    for i, w in enumerate([8, 22, 14, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
